# Import {{{
import re
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from lib.utils import get_file_path
# }}}

logger = logging.getLogger(__name__)


def make_xls(file_name, worksheet_name, worksheet_headers, rows):
    if not rows:
        return

    # Create a new workbook.
    workbook = Workbook()

    # Add a new sheet.
    sheet = workbook.active
    sheet.title = worksheet_name

    # Add headers
    sheet.append([header.capitalize() for header in worksheet_headers])

    # Colum widths will be stored here
    column_widths = [10 for _ in worksheet_headers]
    no_newlie_re = re.compile(r'^[^\n]*$')
    # Add rows
    for row in rows:
        sheet_row = []
        for index, header in enumerate(worksheet_headers):
            row_value = row[header]
            # Don't process empty values.
            if row_value is None:
                sheet_row.append(None)
                continue

            # Calculate row length.
            if no_newlie_re.match(row_value):
                row_length = len(row_value)
            else:
                row_length = len(max(row_value.split(u"\n"), key=len))

            # Adjust column width according to text length.
            if row_length > column_widths[index]:
                column_widths[index] = row_length

            sheet_row.append(row_value)

        try:
            sheet.append(sheet_row)
        except ValueError as e:
            logger.error(f'Error appending row to sheet: {e}')

    # Set column dimensions.
    for index in range(len(worksheet_headers)):
        # Worksheet indices start from 1.
        column_letter = get_column_letter(index + 1)
        column_dimensions = sheet.column_dimensions[column_letter]

        column_dimensions.width = column_widths[index]

    # Write workbook.
    workbook_path = get_file_path('var', f'{file_name}.xlsx')
    workbook.save(workbook_path)


def open_workbook(*file_name):
    if not file_name:
        return None, 'Please provide a correct xls file.'

    workbook, error = None, None
    try:
        workbook_path = get_file_path(*file_name)
        workbook = load_workbook(workbook_path)
    except (IOError, InvalidFileException) as e:
        message = e.strerror if hasattr(e, 'strerror') else e.message
        error = 'Error: {0}'.format(message)

    return workbook, error


def save_workbook(workbook, *file_name):
    workbook_path = get_file_path(*file_name)
    workbook.save(workbook_path)
